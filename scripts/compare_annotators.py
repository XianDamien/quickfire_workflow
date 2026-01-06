# -*- coding: utf-8 -*-
"""
比较不同 Annotator 的标注效果

用法:
    python scripts/compare_annotators.py \
        --archive-batch Zoe41900_2025-09-08 \
        --annotators qwen3-max gemini-relay \
        --output-dir archive/Zoe41900_2025-09-08/comparison_reports
"""

import os
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed

# 添加项目根目录到路径
sys.path.insert(0, str(Path(__file__).parent.parent))

from scripts.annotators import get_annotator
from scripts.annotators.base import AnnotatorOutput
from scripts.common.runs import new_run_id, ensure_run_dir


def load_env():
    """加载环境变量"""
    env_path = Path(__file__).parent / ".env"
    if env_path.exists():
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, value = line.split("=", 1)
                    os.environ.setdefault(key.strip(), value.strip())


def get_students(archive_batch: str) -> List[str]:
    """获取 archive batch 中的所有学生"""
    archive_dir = Path(__file__).parent.parent / "archive" / archive_batch
    metadata_path = archive_dir / "metadata.json"

    if metadata_path.exists():
        with open(metadata_path) as f:
            metadata = json.load(f)
        return [item["student"] for item in metadata.get("items", [])]

    # 回退：扫描目录
    students = []
    for p in archive_dir.iterdir():
        if p.is_dir() and not p.name.startswith(".") and p.name != "reports":
            students.append(p.name)
    return sorted(students)


def run_annotator(
    annotator_name: str,
    archive_batch: str,
    student: str,
    force_relay: Optional[bool] = None,
    verbose: bool = False
) -> Dict[str, Any]:
    """
    运行单个 annotator 处理单个学生

    Returns:
        包含结果的字典
    """
    start_time = datetime.now()

    try:
        # 创建 annotator
        if annotator_name == "gemini-relay":
            annotator = get_annotator("gemini-2.5-pro", force_relay=True)
            display_name = "gemini-2.5-pro (relay)"
        elif annotator_name == "gemini-official":
            annotator = get_annotator("gemini-2.5-pro", force_relay=False)
            display_name = "gemini-2.5-pro (official)"
        else:
            annotator = get_annotator(annotator_name)
            display_name = annotator_name

        # 创建 run 目录
        run_id = new_run_id()
        run_dir = ensure_run_dir(archive_batch, student, display_name.replace(" ", "_"), run_id)

        # 运行标注
        result: AnnotatorOutput = annotator.run_archive_student(
            archive_batch=archive_batch,
            student_name=student,
            run_dir=run_dir,
            force=True,
            verbose=verbose
        )

        end_time = datetime.now()
        elapsed_ms = (end_time - start_time).total_seconds() * 1000

        return {
            "student": student,
            "annotator": display_name,
            "success": result.success,
            "grade": result.final_grade if result.success else None,
            "mistake_count": result.mistake_count if result.success else None,
            "annotation_count": len(result.annotations) if result.success else 0,
            "api_response_time_ms": result.response_time_ms,
            "total_time_ms": elapsed_ms,
            "model": result.model,
            "run_id": result.run_id,
            "error": result.error
        }

    except Exception as e:
        end_time = datetime.now()
        elapsed_ms = (end_time - start_time).total_seconds() * 1000
        return {
            "student": student,
            "annotator": annotator_name,
            "success": False,
            "grade": None,
            "mistake_count": None,
            "annotation_count": 0,
            "api_response_time_ms": None,
            "total_time_ms": elapsed_ms,
            "model": None,
            "run_id": None,
            "error": str(e)
        }


def generate_comparison_report(
    results: Dict[str, List[Dict]],
    archive_batch: str,
    output_dir: Path
) -> Dict[str, Any]:
    """生成对比报告"""

    report = {
        "test_info": {
            "batch": archive_batch,
            "annotators": list(results.keys()),
            "timestamp": datetime.now().isoformat(),
        },
        "results_by_annotator": results,
        "comparison": []
    }

    # 获取所有学生
    students = set()
    for annotator_results in results.values():
        for r in annotator_results:
            students.add(r["student"])

    # 按学生对比
    for student in sorted(students):
        student_comparison = {"student": student}

        for annotator_name, annotator_results in results.items():
            student_result = next(
                (r for r in annotator_results if r["student"] == student),
                None
            )
            if student_result:
                student_comparison[annotator_name] = {
                    "grade": student_result["grade"],
                    "mistake_count": student_result.get("mistake_count", {}).get("errors", "N/A"),
                    "annotation_count": student_result["annotation_count"],
                    "response_time_ms": student_result["api_response_time_ms"],
                    "success": student_result["success"],
                    "error": student_result.get("error")
                }

        report["comparison"].append(student_comparison)

    # 汇总统计
    summary = {}
    for annotator_name, annotator_results in results.items():
        successful = [r for r in annotator_results if r["success"]]
        summary[annotator_name] = {
            "total": len(annotator_results),
            "successful": len(successful),
            "failed": len(annotator_results) - len(successful),
            "grade_distribution": {},
            "avg_response_time_ms": None,
            "avg_annotation_count": None
        }

        if successful:
            # 成绩分布
            grades = [r["grade"] for r in successful if r["grade"]]
            for grade in ["A", "B", "C"]:
                summary[annotator_name]["grade_distribution"][grade] = grades.count(grade)

            # 平均响应时间
            response_times = [r["api_response_time_ms"] for r in successful if r["api_response_time_ms"]]
            if response_times:
                summary[annotator_name]["avg_response_time_ms"] = sum(response_times) / len(response_times)

            # 平均标注数量
            annotation_counts = [r["annotation_count"] for r in successful]
            summary[annotator_name]["avg_annotation_count"] = sum(annotation_counts) / len(annotation_counts)

    report["summary"] = summary

    return report


def generate_markdown_report(report: Dict[str, Any], output_path: Path):
    """生成 Markdown 格式报告"""
    lines = []

    lines.append(f"# Annotator 对比报告")
    lines.append("")
    lines.append(f"**批次**: {report['test_info']['batch']}")
    lines.append(f"**时间**: {report['test_info']['timestamp']}")
    lines.append(f"**对比模型**: {', '.join(report['test_info']['annotators'])}")
    lines.append("")

    # 汇总统计
    lines.append("## 汇总统计")
    lines.append("")

    annotators = list(report["summary"].keys())

    # 表头
    header = "| 指标 | " + " | ".join(annotators) + " |"
    separator = "|---" + "|---" * len(annotators) + "|"
    lines.append(header)
    lines.append(separator)

    # 成功率
    row = "| 成功/总数 |"
    for ann in annotators:
        s = report["summary"][ann]
        row += f" {s['successful']}/{s['total']} |"
    lines.append(row)

    # 成绩分布
    for grade in ["A", "B", "C"]:
        row = f"| 成绩 {grade} |"
        for ann in annotators:
            count = report["summary"][ann]["grade_distribution"].get(grade, 0)
            row += f" {count} |"
        lines.append(row)

    # 平均响应时间
    row = "| 平均响应时间 (ms) |"
    for ann in annotators:
        t = report["summary"][ann]["avg_response_time_ms"]
        row += f" {t:.0f} |" if t else " N/A |"
    lines.append(row)

    # 平均标注数
    row = "| 平均标注数 |"
    for ann in annotators:
        c = report["summary"][ann]["avg_annotation_count"]
        row += f" {c:.1f} |" if c else " N/A |"
    lines.append(row)

    lines.append("")

    # 详细对比
    lines.append("## 学生详细对比")
    lines.append("")

    # 表头
    header_parts = ["| 学生 |"]
    for ann in annotators:
        header_parts.append(f" {ann} 成绩 | 错误数 | 响应时间 |")
    lines.append("".join(header_parts))

    separator_parts = ["|---"]
    for _ in annotators:
        separator_parts.append("|:---:|:---:|---:")
    separator_parts.append("|")
    lines.append("".join(separator_parts))

    # 每个学生的数据
    for comp in report["comparison"]:
        row = f"| {comp['student']} |"
        for ann in annotators:
            data = comp.get(ann, {})
            if data.get("success"):
                grade = data.get("grade", "N/A")
                errors = data.get("mistake_count", "N/A")
                time_ms = data.get("response_time_ms")
                time_str = f"{time_ms:.0f}ms" if time_ms else "N/A"
                row += f" {grade} | {errors} | {time_str} |"
            else:
                error = data.get("error", "Failed")[:20]
                row += f" ❌ | - | {error}... |"
        lines.append(row)

    lines.append("")

    # 一致性分析
    lines.append("## 一致性分析")
    lines.append("")

    if len(annotators) == 2:
        ann1, ann2 = annotators
        matches = 0
        mismatches = []

        for comp in report["comparison"]:
            data1 = comp.get(ann1, {})
            data2 = comp.get(ann2, {})

            if data1.get("success") and data2.get("success"):
                if data1.get("grade") == data2.get("grade"):
                    matches += 1
                else:
                    mismatches.append({
                        "student": comp["student"],
                        ann1: data1.get("grade"),
                        ann2: data2.get("grade")
                    })

        total = len([c for c in report["comparison"]
                    if c.get(ann1, {}).get("success") and c.get(ann2, {}).get("success")])

        if total > 0:
            lines.append(f"- **成绩一致率**: {matches}/{total} ({matches/total*100:.1f}%)")
            lines.append("")

            if mismatches:
                lines.append("### 成绩不一致的学生")
                lines.append("")
                lines.append(f"| 学生 | {ann1} | {ann2} |")
                lines.append("|---|:---:|:---:|")
                for m in mismatches:
                    lines.append(f"| {m['student']} | {m[ann1]} | {m[ann2]} |")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def generate_excel_report(report: Dict[str, Any], output_path: Path):
    """生成 Excel 格式报告"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("Warning: openpyxl 未安装，跳过 Excel 报告生成")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "对比报告"

    annotators = list(report["summary"].keys())

    # 标题
    ws["A1"] = "Annotator 对比报告"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"批次: {report['test_info']['batch']}"
    ws["A3"] = f"时间: {report['test_info']['timestamp']}"

    # 汇总表
    row = 5
    ws.cell(row, 1, "汇总统计").font = Font(bold=True)
    row += 1

    headers = ["指标"] + annotators
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row, col, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="CCCCCC")

    row += 1
    metrics = [
        ("成功/总数", lambda s: f"{s['successful']}/{s['total']}"),
        ("成绩 A", lambda s: s["grade_distribution"].get("A", 0)),
        ("成绩 B", lambda s: s["grade_distribution"].get("B", 0)),
        ("成绩 C", lambda s: s["grade_distribution"].get("C", 0)),
        ("平均响应时间 (ms)", lambda s: f"{s['avg_response_time_ms']:.0f}" if s['avg_response_time_ms'] else "N/A"),
        ("平均标注数", lambda s: f"{s['avg_annotation_count']:.1f}" if s['avg_annotation_count'] else "N/A"),
    ]

    for metric_name, metric_fn in metrics:
        ws.cell(row, 1, metric_name)
        for col, ann in enumerate(annotators, 2):
            ws.cell(row, col, metric_fn(report["summary"][ann]))
        row += 1

    # 详细数据
    row += 2
    ws.cell(row, 1, "学生详细对比").font = Font(bold=True)
    row += 1

    headers = ["学生"]
    for ann in annotators:
        headers.extend([f"{ann} 成绩", "错误数", "响应时间(ms)"])

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row, col, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="CCCCCC")

    row += 1
    for comp in report["comparison"]:
        col = 1
        ws.cell(row, col, comp["student"])
        col += 1

        for ann in annotators:
            data = comp.get(ann, {})
            if data.get("success"):
                ws.cell(row, col, data.get("grade", "N/A"))
                ws.cell(row, col + 1, data.get("mistake_count", "N/A"))
                ws.cell(row, col + 2, data.get("response_time_ms"))
            else:
                ws.cell(row, col, "❌")
                ws.cell(row, col + 1, "-")
                ws.cell(row, col + 2, data.get("error", "")[:30])
            col += 3

        row += 1

    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="比较不同 Annotator 的标注效果")
    parser.add_argument("--archive-batch", required=True, help="Archive batch 名称")
    parser.add_argument("--annotators", nargs="+", default=["qwen3-max", "gemini-relay"],
                       help="要比较的 annotator 列表")
    parser.add_argument("--students", nargs="*", help="指定学生列表，默认处理所有学生")
    parser.add_argument("--output-dir", help="输出目录")
    parser.add_argument("--parallel", type=int, default=1, help="并行处理的线程数")
    parser.add_argument("--verbose", action="store_true", help="详细输出")

    args = parser.parse_args()

    # 加载环境变量
    load_env()

    # 获取学生列表
    students = args.students or get_students(args.archive_batch)
    print(f"📚 批次: {args.archive_batch}")
    print(f"👥 学生: {', '.join(students)}")
    print(f"🤖 Annotators: {', '.join(args.annotators)}")
    print()

    # 运行所有 annotator
    results = {ann: [] for ann in args.annotators}

    for annotator_name in args.annotators:
        print(f"🚀 运行 {annotator_name}...")

        if args.parallel > 1:
            # 并行处理
            with ThreadPoolExecutor(max_workers=args.parallel) as executor:
                futures = {
                    executor.submit(
                        run_annotator,
                        annotator_name,
                        args.archive_batch,
                        student,
                        verbose=args.verbose
                    ): student
                    for student in students
                }

                for future in as_completed(futures):
                    student = futures[future]
                    try:
                        result = future.result()
                        results[annotator_name].append(result)
                        status = "✅" if result["success"] else "❌"
                        grade = result.get("grade", "N/A")
                        print(f"  {status} {student}: {grade}")
                    except Exception as e:
                        print(f"  ❌ {student}: {e}")
        else:
            # 串行处理
            for student in students:
                print(f"  处理 {student}...", end=" ", flush=True)
                result = run_annotator(
                    annotator_name,
                    args.archive_batch,
                    student,
                    verbose=args.verbose
                )
                results[annotator_name].append(result)

                if result["success"]:
                    print(f"✅ {result['grade']} (错误: {result['mistake_count'].get('errors', 'N/A')})")
                else:
                    print(f"❌ {result['error'][:50]}")

        print()

    # 生成报告
    output_dir = Path(args.output_dir) if args.output_dir else (
        Path(__file__).parent.parent / "archive" / args.archive_batch / "comparison_reports"
    )
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    print("📊 生成报告...")

    # JSON 报告
    report = generate_comparison_report(results, args.archive_batch, output_dir)
    json_path = output_dir / f"comparison_{timestamp}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"  📄 JSON: {json_path}")

    # Markdown 报告
    md_path = output_dir / f"comparison_{timestamp}.md"
    generate_markdown_report(report, md_path)
    print(f"  📝 Markdown: {md_path}")

    # Excel 报告
    xlsx_path = output_dir / f"comparison_{timestamp}.xlsx"
    generate_excel_report(report, xlsx_path)
    print(f"  📊 Excel: {xlsx_path}")

    print()
    print("✨ 完成!")

    # 打印简要汇总
    print()
    print("=" * 60)
    print("汇总")
    print("=" * 60)

    for ann, summary in report["summary"].items():
        print(f"\n{ann}:")
        print(f"  成功率: {summary['successful']}/{summary['total']}")
        print(f"  成绩分布: A={summary['grade_distribution'].get('A', 0)}, "
              f"B={summary['grade_distribution'].get('B', 0)}, "
              f"C={summary['grade_distribution'].get('C', 0)}")
        if summary['avg_response_time_ms']:
            print(f"  平均响应时间: {summary['avg_response_time_ms']:.0f}ms")


if __name__ == "__main__":
    main()
