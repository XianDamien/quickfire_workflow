#!/usr/bin/env python3
"""
ASR vs 音频方案对比工具

功能：
1. 对比两种方案的测试结果
2. 生成详细的 Excel 对比报告
3. 提供客观的统计数据
"""

import json
import sys
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def load_batch_manifest(run_dir: Path) -> Optional[dict]:
    """加载 batch_manifest.json"""
    manifest_file = run_dir / "batch_manifest.json"
    if manifest_file.exists():
        with open(manifest_file) as f:
            return json.load(f)
    return None


def extract_run_info(manifest: dict, run_dir_name: str) -> dict:
    """提取运行信息"""
    # 推断 mode（如果未明确指定）
    mode = manifest.get("mode", "")
    if not mode:
        if "audio" in run_dir_name.lower():
            mode = "audio"
        else:
            mode = "asr"

    # 兼容新旧格式的 students_count
    statistics = manifest.get("statistics", {})
    student_results = manifest.get("student_results", [])
    students_count = statistics.get("students_count") or manifest.get("students_count") or len(student_results)
    success_count = statistics.get("success_count")
    fail_count = statistics.get("fail_count")

    if success_count is None or fail_count is None:
        success_count = sum(1 for s in student_results if s.get("status") == "success")
        fail_count = sum(1 for s in student_results if s.get("status") != "success")

    grade_distribution = statistics.get("grade_distribution")
    if not grade_distribution:
        grade_distribution = {"A": 0, "B": 0, "C": 0}
        for s in student_results:
            grade = s.get("grade")
            if grade in grade_distribution:
                grade_distribution[grade] += 1

    return {
        "run_id": manifest.get("run_id", ""),
        "mode": mode,
        "model": manifest.get("model", ""),
        "success_count": success_count,
        "fail_count": fail_count,
        "students_count": students_count,
        "grade_distribution": grade_distribution,
        "prompt_tokens": manifest.get("token_usage", {}).get("prompt_tokens", 0),
        "thoughts_tokens": manifest.get("token_usage", {}).get("thoughts_tokens", 0),
        "candidates_tokens": manifest.get("token_usage", {}).get("candidates_tokens", 0),
        "total_tokens": manifest.get("token_usage", {}).get("total_tokens", 0),
        "processing_time": manifest.get("timing", {}).get("api_processing_time_seconds", 0),
        "student_results": student_results
    }


def compare_runs(batch_dir: Path, run_ids: list[str]) -> dict:
    """对比多个运行结果"""
    results = []

    batch_runs_dir = batch_dir / "_batch_runs"
    if not batch_runs_dir.exists():
        print(f"错误: {batch_runs_dir} 不存在")
        return {}

    for run_id in run_ids:
        # 查找匹配的目录
        matching_dirs = [d for d in batch_runs_dir.iterdir() if d.is_dir() and run_id in d.name]

        if not matching_dirs:
            print(f"警告: 未找到 run_id={run_id} 的目录")
            continue

        run_dir = matching_dirs[0]
        manifest = load_batch_manifest(run_dir)

        if not manifest:
            print(f"警告: {run_dir.name} 没有 batch_manifest.json")
            continue

        info = extract_run_info(manifest, run_dir.name)
        info["run_dir"] = run_dir.name
        results.append(info)

    return {
        "batch": batch_dir.name,
        "runs": results
    }


def create_comparison_excel(comparison_data: dict, output_file: Path):
    """创建对比 Excel 报告"""
    wb = openpyxl.Workbook()

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: 核心指标
    ws1 = wb.active
    ws1.title = "核心指标"

    headers = ["指标"] + [f"测试 {i+1}\n({run['mode']})" for i, run in enumerate(comparison_data["runs"])]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    def format_success_rate(run: dict) -> str:
        total = run.get("students_count", 0)
        success = run.get("success_count", 0)
        if total <= 0:
            return "N/A"
        pct = success / total * 100
        return f"{success}/{total} ({pct:.1f}%)"

    def format_grade_distribution(run: dict) -> str:
        dist = run.get("grade_distribution", {})
        return f"A:{dist.get('A', 0)} B:{dist.get('B', 0)} C:{dist.get('C', 0)}"

    def format_token_distribution(run: dict) -> str:
        total = run.get("total_tokens", 0)
        if total <= 0:
            return "N/A"
        prompt_pct = run.get("prompt_tokens", 0) / total * 100
        thoughts_pct = run.get("thoughts_tokens", 0) / total * 100
        candidates_pct = run.get("candidates_tokens", 0) / total * 100
        return f"P:{prompt_pct:.1f}% T:{thoughts_pct:.1f}% C:{candidates_pct:.1f}%"

    def format_avg_errors(run: dict) -> str:
        errors = []
        for result in run.get("student_results", []):
            if result.get("status") == "success":
                errors.append(result.get("mistake_count", {}).get("errors", 0))
        if not errors:
            return "0"
        return f"{sum(errors) / len(errors):.2f}"

    def format_failed_students(run: dict) -> str:
        failed_students = [r.get("student") for r in run.get("student_results", []) if r.get("status") != "success"]
        failed_students = [s for s in failed_students if s]
        if not failed_students:
            return "0"
        return f"{len(failed_students)} ({', '.join(failed_students)})"

    metrics = [
        ("批次ID", lambda _: comparison_data["batch"]),
        ("Run ID", "run_id"),
        ("模型", "model"),
        ("方案", "mode"),
        ("学生总数", "students_count"),
        ("成功率", format_success_rate),
        ("成绩分布", format_grade_distribution),
        ("Total Tokens", "total_tokens"),
        ("Token分布", format_token_distribution),
        ("处理时间 (秒)", lambda r: f"{r.get('processing_time', 0):.1f}"),
        ("平均错误数", format_avg_errors),
        ("失败学生数", format_failed_students),
    ]

    row = 2
    for metric_name, metric_key in metrics:
        ws1.cell(row=row, column=1, value=metric_name).border = thin_border
        for col, run in enumerate(comparison_data["runs"], 2):
            value = metric_key(run) if callable(metric_key) else run.get(metric_key, "")
            cell = ws1.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        row += 1

    ws1.column_dimensions["A"].width = 18
    for col in range(2, len(headers) + 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

    # Sheet 2: 错误详情
    ws2 = wb.create_sheet("错误详情")
    error_headers = ["批次", "Run ID", "学生", "题号", "时间戳", "题目", "期望答案", "实际回答", "错误类型", "成绩"]
    for col, header in enumerate(error_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    batch_dir = output_file.parent
    row = 2
    for run in comparison_data["runs"]:
        model = run.get("model", "")
        run_id = run.get("run_id", "")
        mode = run.get("mode", "")
        annotator_candidates = []
        if mode == "audio":
            annotator_candidates = [f"{model}.audio", f"{model}_audio", model]
        else:
            annotator_candidates = [model]

        for result in run.get("student_results", []):
            student_name = result.get("student")
            if not student_name:
                continue

            ann_path = None
            for annotator in annotator_candidates:
                candidate = batch_dir / student_name / "runs" / annotator / run_id / "4_llm_annotation.json"
                if candidate.exists():
                    ann_path = candidate
                    break

            if not ann_path or not ann_path.exists():
                continue

            with open(ann_path, "r", encoding="utf-8") as f:
                ann_data = json.load(f)

            final_grade = ann_data.get("final_grade_suggestion")
            annotations = ann_data.get("annotations", [])

            for ann in annotations:
                utterance = ann.get("related_student_utterance", {}) if isinstance(ann, dict) else {}
                issue_type = utterance.get("issue_type") or ann.get("issue_type")
                if not issue_type:
                    continue

                detected_text = utterance.get("detected_text")
                if detected_text is None:
                    detected_text = ann.get("detected_answer") or ann.get("detected_text")

                ws2.cell(row=row, column=1, value=comparison_data["batch"]).border = thin_border
                ws2.cell(row=row, column=2, value=run_id).border = thin_border
                ws2.cell(row=row, column=3, value=student_name).border = thin_border
                ws2.cell(row=row, column=4, value=ann.get("card_index")).border = thin_border
                ws2.cell(row=row, column=5, value=ann.get("card_timestamp")).border = thin_border
                ws2.cell(row=row, column=6, value=ann.get("question")).border = thin_border
                ws2.cell(row=row, column=7, value=ann.get("expected_answer")).border = thin_border
                ws2.cell(row=row, column=8, value=detected_text).border = thin_border
                ws2.cell(row=row, column=9, value=issue_type).border = thin_border
                ws2.cell(row=row, column=10, value=final_grade).border = thin_border
                row += 1

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 8
    ws2.column_dimensions["E"].width = 12
    ws2.column_dimensions["F"].width = 26
    ws2.column_dimensions["G"].width = 22
    ws2.column_dimensions["H"].width = 22
    ws2.column_dimensions["I"].width = 14
    ws2.column_dimensions["J"].width = 8

    # 保存
    wb.save(output_file)
    print(f"\n已生成对比报告: {output_file}")


def main():
    if len(sys.argv) < 3:
        print("用法: uv run python scripts/compare_asr_audio.py <batch_id> <run_id1> [run_id2] [run_id3] ...")
        print("\n示例:")
        print("  uv run python scripts/compare_asr_audio.py Zoe61330_2025-12-15 002933 110826 192906")
        print("\n说明:")
        print("  - batch_id: 班级目录名（如 Zoe61330_2025-12-15）")
        print("  - run_id: 运行 ID 的关键部分（如 002933 匹配 20260106_002933_eb28926）")
        sys.exit(1)

    batch_id = sys.argv[1]
    run_ids = sys.argv[2:]

    archive_dir = Path(__file__).parent.parent / "archive"
    batch_dir = archive_dir / batch_id

    if not batch_dir.exists():
        print(f"错误: 批次目录不存在 {batch_dir}")
        sys.exit(1)

    print(f"对比批次: {batch_id}")
    print(f"运行数: {len(run_ids)}")

    # 对比
    comparison_data = compare_runs(batch_dir, run_ids)

    if not comparison_data.get("runs"):
        print("错误: 未找到有效的运行数据")
        sys.exit(1)

    # 生成报告
    output_file = batch_dir / f"comparison_report_{'-'.join(run_ids)}.xlsx"
    create_comparison_excel(comparison_data, output_file)

    # 打印摘要
    print("\n" + "=" * 60)
    print("对比摘要")
    print("=" * 60)

    for i, run in enumerate(comparison_data["runs"], 1):
        print(f"\n测试 {i} ({run['mode']}):")
        print(f"  Run ID: {run['run_id']}")

        # 计算成功率
        if run['students_count'] > 0:
            success_rate = run['success_count'] / run['students_count'] * 100
            print(f"  成功率: {run['success_count']}/{run['students_count']} ({success_rate:.1f}%)")
        else:
            print(f"  成功率: {run['success_count']}/{run['students_count']} (N/A)")

        print(f"  Total Tokens: {run['total_tokens']:,}")

        # 计算 thoughts token 占比
        if run['total_tokens'] > 0:
            thoughts_pct = run['thoughts_tokens'] / run['total_tokens'] * 100
            print(f"  Thoughts Tokens: {run['thoughts_tokens']:,} ({thoughts_pct:.1f}%)")
        else:
            print(f"  Thoughts Tokens: {run['thoughts_tokens']:,}")

        print(f"  处理时间: {run['processing_time']:.1f}s")


if __name__ == "__main__":
    main()
