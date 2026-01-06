#!/usr/bin/env python3
"""
整合所有 archive 数据到一个 xlsx 文件

功能：
1. 遍历所有 archive 目录
2. 提取每个学生最新的 LLM annotation 结果
3. 生成成绩总表和错误详情 sheet
4. 识别未测试的数据集
"""

import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


ARCHIVE_DIR = Path("/Users/damien/Desktop/Venture/quickfire_workflow/archive")
OUTPUT_FILE = ARCHIVE_DIR / "consolidated_grades.xlsx"


def parse_class_info(folder_name: str) -> dict:
    """解析班级文件夹名称，提取老师名和班级代码"""
    # 格式: Zoe61330_2025-12-15 -> teacher=Zoe, class_code=61330, date=2025-12-15
    match = re.match(r"([A-Za-z]+)(\d+)_(\d{4}-\d{2}-\d{2})", folder_name)
    if match:
        teacher = match.group(1)
        class_num = match.group(2)
        date = match.group(3)
        # 格式化班级代码: Zoe-6-1330 from Zoe61330
        if len(class_num) >= 4:
            class_code = f"{teacher}-{class_num[0]}-{class_num[1:]}"
        else:
            class_code = f"{teacher}-{class_num}"
        return {
            "teacher": teacher,
            "class_num": class_num,
            "class_code": class_code,
            "date": date
        }
    return {}


def find_latest_annotation(student_dir: Path) -> Optional[tuple[Path, datetime]]:
    """找到学生目录下最新的 4_llm_annotation.json"""
    runs_dir = student_dir / "runs"
    if not runs_dir.exists():
        # 检查旧格式（直接在学生目录下）
        direct_annotation = student_dir / "4_llm_annotation.json"
        if direct_annotation.exists():
            mtime = datetime.fromtimestamp(direct_annotation.stat().st_mtime)
            return direct_annotation, mtime
        return None

    latest_file = None
    latest_time = None

    # 遍历所有模型目录
    for model_dir in runs_dir.iterdir():
        if not model_dir.is_dir():
            continue
        # 遍历所有运行目录
        for run_dir in model_dir.iterdir():
            if not run_dir.is_dir():
                continue
            annotation_file = run_dir / "4_llm_annotation.json"
            if annotation_file.exists():
                # 从目录名解析时间戳: 20251230_005955_4d4ce29
                try:
                    timestamp_str = run_dir.name[:15]  # 20251230_005955
                    run_time = datetime.strptime(timestamp_str, "%Y%m%d_%H%M%S")
                    if latest_time is None or run_time > latest_time:
                        latest_time = run_time
                        latest_file = annotation_file
                except ValueError:
                    # 如果解析失败，使用文件修改时间
                    mtime = datetime.fromtimestamp(annotation_file.stat().st_mtime)
                    if latest_time is None or mtime > latest_time:
                        latest_time = mtime
                        latest_file = annotation_file

    return (latest_file, latest_time) if latest_file else None


def load_metadata(batch_dir: Path) -> dict:
    """加载 metadata.json"""
    metadata_file = batch_dir / "metadata.json"
    if metadata_file.exists():
        with open(metadata_file) as f:
            return json.load(f)
    return {}


def extract_d_number(progress: str) -> str:
    """从 progress 字段提取 D 几

    格式示例:
    - R1-27-D2 -> D2
    - R3-14-D4 -> D4
    - 130-26-EC -> EC (高中英语)
    """
    if not progress:
        return ""

    # 格式1: R{round}-{num}-D{level} -> 直接提取 D{level}
    match = re.search(r"-D(\d+)$", progress)
    if match:
        return f"D{match.group(1)}"

    # 格式2: {num}-{sub}-EC -> EC (高中英语)
    if progress.endswith("-EC"):
        return "EC"

    # 格式3: 其他格式，返回原始 progress
    return progress


def process_all_batches():
    """处理所有批次目录"""
    grades_data = []
    errors_data = []
    untested_data = []

    # 遍历 archive 目录
    for batch_dir in sorted(ARCHIVE_DIR.iterdir()):
        # 跳过非目录、zip 文件、funasr 测试目录
        if not batch_dir.is_dir():
            continue
        if batch_dir.name.endswith(".zip"):
            continue
        if "funasr" in batch_dir.name.lower():
            continue
        if batch_dir.name.startswith("."):
            continue

        class_info = parse_class_info(batch_dir.name)
        if not class_info:
            print(f"无法解析目录名: {batch_dir.name}")
            continue

        metadata = load_metadata(batch_dir)
        progress = metadata.get("progress", "")
        d_number = extract_d_number(progress)

        print(f"处理: {batch_dir.name} (D={d_number}, progress={progress})")

        # 遍历学生目录
        for student_dir in sorted(batch_dir.iterdir()):
            if not student_dir.is_dir():
                continue
            if student_dir.name.startswith("."):
                continue
            if student_dir.name in ["reports", "_shared_context"]:
                continue

            student_name = student_dir.name

            # 查找最新的 annotation
            result = find_latest_annotation(student_dir)

            if result is None:
                # 未测试
                untested_data.append({
                    "batch": batch_dir.name,
                    "teacher": class_info["teacher"],
                    "date": class_info["date"],
                    "class_code": class_info["class_code"],
                    "student": student_name,
                    "reason": "无 4_llm_annotation.json"
                })
                continue

            annotation_file, run_time = result

            try:
                with open(annotation_file) as f:
                    annotation = json.load(f)
            except Exception as e:
                print(f"  读取失败: {annotation_file}: {e}")
                continue

            # 提取成绩
            grade = annotation.get("final_grade_suggestion", "")
            mistake_count = annotation.get("mistake_count", {})
            error_count = mistake_count.get("errors", 0) if isinstance(mistake_count, dict) else 0

            # 添加成绩记录
            grades_data.append({
                "teacher": class_info["teacher"],
                "date": class_info["date"],
                "d_number": d_number,
                "class_code": class_info["class_code"],
                "student": student_name,
                "grade": grade,
                "error_count": error_count,
                "run_time": run_time.strftime("%Y-%m-%d %H:%M"),
                "batch": batch_dir.name
            })

            # 提取错误详情
            annotations = annotation.get("annotations", [])
            for item in annotations:
                issue_type = None
                utterance = item.get("related_student_utterance", {})
                if isinstance(utterance, dict):
                    issue_type = utterance.get("issue_type")

                if issue_type:
                    errors_data.append({
                        "batch": batch_dir.name,
                        "teacher": class_info["teacher"],
                        "date": class_info["date"],
                        "class_code": class_info["class_code"],
                        "student": student_name,
                        "card_index": item.get("card_index", ""),
                        "question": item.get("question", ""),
                        "expected_answer": item.get("expected_answer", ""),
                        "detected_text": utterance.get("detected_text", "") if isinstance(utterance, dict) else "",
                        "issue_type": issue_type
                    })

    return grades_data, errors_data, untested_data


def create_excel(grades_data, errors_data, untested_data):
    """创建 Excel 文件"""
    wb = openpyxl.Workbook()

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: 成绩总表
    ws1 = wb.active
    ws1.title = "成绩总表"

    headers1 = ["老师", "日期", "D级别", "班级", "学生", "成绩", "错误数", "处理时间", "批次ID"]
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row_idx, data in enumerate(grades_data, 2):
        ws1.cell(row=row_idx, column=1, value=data["teacher"]).border = thin_border
        ws1.cell(row=row_idx, column=2, value=data["date"]).border = thin_border
        ws1.cell(row=row_idx, column=3, value=data["d_number"]).border = thin_border
        ws1.cell(row=row_idx, column=4, value=data["class_code"]).border = thin_border
        ws1.cell(row=row_idx, column=5, value=data["student"]).border = thin_border
        grade_cell = ws1.cell(row=row_idx, column=6, value=data["grade"])
        grade_cell.border = thin_border
        grade_cell.alignment = Alignment(horizontal="center")
        # 根据成绩着色
        if data["grade"] == "A":
            grade_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif data["grade"] == "C":
            grade_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        ws1.cell(row=row_idx, column=7, value=data["error_count"]).border = thin_border
        ws1.cell(row=row_idx, column=8, value=data["run_time"]).border = thin_border
        ws1.cell(row=row_idx, column=9, value=data["batch"]).border = thin_border

    # 调整列宽
    for col in range(1, len(headers1) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 15

    # Sheet 2: 错误详情
    ws2 = wb.create_sheet("错误详情")

    headers2 = ["批次", "老师", "日期", "班级", "学生", "题号", "题目", "期望答案", "学生回答", "错误类型"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row_idx, data in enumerate(errors_data, 2):
        ws2.cell(row=row_idx, column=1, value=data["batch"]).border = thin_border
        ws2.cell(row=row_idx, column=2, value=data["teacher"]).border = thin_border
        ws2.cell(row=row_idx, column=3, value=data["date"]).border = thin_border
        ws2.cell(row=row_idx, column=4, value=data["class_code"]).border = thin_border
        ws2.cell(row=row_idx, column=5, value=data["student"]).border = thin_border
        ws2.cell(row=row_idx, column=6, value=data["card_index"]).border = thin_border
        ws2.cell(row=row_idx, column=7, value=data["question"]).border = thin_border
        ws2.cell(row=row_idx, column=8, value=data["expected_answer"]).border = thin_border
        detected = data["detected_text"] or "(无回答)"
        ws2.cell(row=row_idx, column=9, value=detected).border = thin_border
        issue_cell = ws2.cell(row=row_idx, column=10, value=data["issue_type"])
        issue_cell.border = thin_border
        # 根据错误类型着色
        if data["issue_type"] == "NO_ANSWER":
            issue_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif data["issue_type"] == "MEANING_ERROR":
            issue_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for col in range(1, len(headers2) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 15
    ws2.column_dimensions["H"].width = 25  # 期望答案列加宽

    # Sheet 3: 未测试数据
    ws3 = wb.create_sheet("未测试数据")

    headers3 = ["批次", "老师", "日期", "班级", "学生", "原因"]
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row_idx, data in enumerate(untested_data, 2):
        ws3.cell(row=row_idx, column=1, value=data["batch"]).border = thin_border
        ws3.cell(row=row_idx, column=2, value=data["teacher"]).border = thin_border
        ws3.cell(row=row_idx, column=3, value=data["date"]).border = thin_border
        ws3.cell(row=row_idx, column=4, value=data["class_code"]).border = thin_border
        ws3.cell(row=row_idx, column=5, value=data["student"]).border = thin_border
        ws3.cell(row=row_idx, column=6, value=data["reason"]).border = thin_border

    for col in range(1, len(headers3) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 20

    # 保存
    wb.save(OUTPUT_FILE)
    print(f"\n已保存到: {OUTPUT_FILE}")

    return {
        "total_students": len(grades_data),
        "total_errors": len(errors_data),
        "untested_students": len(untested_data)
    }


def main():
    print("=" * 60)
    print("整合 Archive 成绩数据")
    print("=" * 60)

    grades_data, errors_data, untested_data = process_all_batches()

    print(f"\n统计:")
    print(f"  已测试学生: {len(grades_data)}")
    print(f"  错误记录: {len(errors_data)}")
    print(f"  未测试学生: {len(untested_data)}")

    stats = create_excel(grades_data, errors_data, untested_data)

    # 打印未测试汇总
    if untested_data:
        print(f"\n未测试数据集:")
        batches = {}
        for item in untested_data:
            batch = item["batch"]
            if batch not in batches:
                batches[batch] = []
            batches[batch].append(item["student"])

        for batch, students in sorted(batches.items()):
            print(f"  {batch}: {', '.join(students)}")


if __name__ == "__main__":
    main()
